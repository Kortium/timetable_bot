import re
from datetime import datetime, timedelta
from enum import Enum

import openpyxl


class DocType(Enum):
    PROFESSOR = 1
    STUDENT = 2


YEAR = 2025  # Определяем константу для года, который будет использоваться в коде


# Функция для извлечения инициалов из названия предмета
def extract_initials(subject_name):
    # Список союзов и предлогов, которые будут использоваться для фильтрации
    conjunctions_and_prepositions = ["и", "или", "но", "да", "что", "как", "на", "под", "за", "с", "к", "в", "до", "из", "у", "при"]

    words = re.split(r'[ -]+', subject_name)  # Разделяем название предмета на слова по пробелам и дефисам
    initials = []
    if len(words) > 1:
        for word in words:
            # Если слово является союзом или предлогом, добавляем его в инициалы как есть
            if word.lower() in conjunctions_and_prepositions:
                initials.append(word.lower())
            else:
                # В противном случае добавляем только первую букву слова, сделав её заглавной
                initials.append(word[0].upper())
    else:
        if len(subject_name) < 12:
            initials.append(subject_name)
        else:
            initials.append(subject_name[0:5])

    return "".join(initials)  # Возвращаем инициалы, объединённые в одну строку


# Функция для сокращения названия группы
def shorten_group(group_name):
    # Преобразуем список групп, оставляя полное название первой группы и сокращая остальные
    if isinstance(group_name, str):
        return group_name
    groups_name = ", ".join(group if i == 0 else group.split("-")[1] if "-" in group else group for i, group in enumerate(group_name))
    return groups_name  # Возвращаем результат


# Функция для извлечения ФИО из текста
def extract_fio(text):
    parts = text.split("преподавателя")  # Разделяем текст по слову "преподавателя"
    if len(parts) > 1:
        return parts[1].strip()  # Если есть результат, возвращаем его, удалив лишние пробелы
    else:
        return None  # Если результат отсутствует, возвращаем None

# Функция для извлечения группы из текста
def extract_group(text):
    parts = text.split("группы")  # Разделяем текст по слову "группы"
    if len(parts) > 1:
        return parts[1].strip()  # Если есть результат, возвращаем его, удалив лишние пробелы
    else:
        return None  # Если результат отсутствует, возвращаем None


def check_type(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    name = extract_fio(sheet['C2'].value)
    if name is not None:
        return DocType.PROFESSOR
    group = extract_group(sheet['C2'].value)
    if group is not None:
        return DocType.STUDENT
    return None


# Функция для извлечения информации о занятиях из текста
def extract_professor_info(text):
    if not text:
        return

    lessons = text.split('---')  # Разделяем текст на отдельные занятия по тройному дефису
    results = []
    for lesson in lessons:
        # Для каждого занятия извлекаем отдельную информацию
        room, subject, lesson_type, group, dates = extract_single_lesson_info(lesson)
        results.append((room, subject, lesson_type, group, dates))
    return results  # Возвращаем список с информацией о занятиях


# Функция для извлечения информации о конкретном занятии
def extract_single_lesson_info(lesson):
    # Ищем информацию о комнате с помощью регулярного выражения
    room_match = re.search(r"ауд\.(каф\.\(-\)|\d+\([А-Яа-я\s]+\)|\d+\(?\d?\)?)", lesson)
    room = room_match.group(1) if room_match else None

    # Ищем тип занятия (ЛР, ЛК, ПЗ)
    lesson_type_match = re.search(r"(ЛР|ЛК|ПЗ)", lesson)
    lesson_type = lesson_type_match.group(1) if lesson_type_match else None

    if room and lesson_type:
        # Если найдены комната и тип, извлекаем название предмета из текста занятия
        subject_start = lesson.find(room) + len(room) + 1
        subject_end = lesson.find(lesson_type) - 2
        subject = lesson[subject_start:subject_end].strip()
    else:
        subject = None

    # Ищем названия групп с помощью регулярного выражения
    group_pattern = r"([А-Я][\dА-Я][А-Я]-\d{3}[А-Яа-я]+-\d{2})"
    groups = re.findall(group_pattern, lesson)

    # Ищем даты занятий
    dates = re.findall(r"(\d{2}\.\d{2}(?:-\d{2}\.\d{2})?)", lesson)
    if dates:
        for i, d in enumerate(dates):
            if '-' not in d:
                dates[i] = f"{d}-{d}"
    else:
        dates = None
    return room, subject, lesson_type, groups, dates  # Возвращаем извлеченные данные о занятии


# Функция для извлечения информации о занятиях из текста
def extract_student_info(text):
    if not text:
        return

    lessons = text.split('---')  # Разделяем текст на отдельные занятия по тройному дефису
    results = []
    for lesson in lessons:
        # Для каждого занятия извлекаем отдельную информацию
        room, subject, lesson_type, professor, dates = extract_single_students_lesson_info(lesson)
        results.append((room, subject, lesson_type, professor, dates))
    return results  # Возвращаем список с информацией о занятиях


def extract_single_students_lesson_info(lesson):
    # Ищем информацию о комнате с помощью регулярного выражения
    room_match = re.search(r"ауд\.(каф\.( *\d*)\(-\)|\d+\([А-Яа-я\s]+|d*\)|Зал [А-Я]\(?\d*\)|\d+\(?\d*\)?)", lesson)
    room = room_match.group(1) if room_match else None

    # Ищем тип занятия (ЛР, ЛК, ПЗ)
    lesson_type_match = re.search(r"(ЛР|ЛК|ПЗ)", lesson)
    lesson_type = lesson_type_match.group(1) if lesson_type_match else None
    if room and lesson_type:
        # Если найдены комната и тип, извлекаем название предмета из текста занятия
        subject_start = lesson.find(room) + len(room) + 1
        subject_end = lesson.find(lesson_type) - 2
        subject = lesson[subject_start:subject_end].strip()
    else:
        subject = None
    # Ищем названия групп с помощью регулярного выражения
    professor_pattern = r"([А-Я][а-я]+ [А-Я]\.+(?:[А-Я]\.)*)"
    professors = re.findall(professor_pattern, lesson)
    if len(professors) > 0:
        professor = professors[0]
    else:
        professor = ''

    # Ищем даты занятий
    dates_match = re.search(r"(\d{2}\.\d{2}(?:-\d{2}\.\d{2})?)", lesson)
    if dates_match:
        dates = dates_match.group(1)
        if "-" not in dates:
            dates = f"{dates}-{dates}"
    else:
        dates = None
    return room, subject, lesson_type, professor, dates  # Возвращаем извлеченные данные о занятии


# Функция для получения дат по определённому дню недели между двумя датами
def get_dates_between(start_date, end_date, weekday):
    current_date = start_date
    while current_date <= end_date:  # Перебираем даты в диапазоне от начальной до конечной
        if current_date.weekday() == weekday:  # Проверяем, соответствует ли текущий день недели заданному
            yield current_date  # Если да, то возвращаем текущую дату
        current_date += timedelta(days=1)  # Переходим к следующему дню

# Функция для проверки, присоединено ли занятие к следующему
def check_if_exercise_joined(exercise, next_cell):
    # Возвращаем True, если в следующей ячейке есть значение и оно содержит информацию о занятии
    return next_cell.value and exercise in extract_professor_info(next_cell.value)

# Функция для получения временного периода занятий на основе буквы столбца
def get_time_period(cell):
    # Словарь, сопоставляющий буквы столбцов с временными промежутками
    time_mapping = {
        'B': ('09:00', '10:30'),
        'C': ('10:45', '12:15'),
        'D': ('13:00', '14:30'),
        'E': ('14:45', '16:15'),
        'F': ('16:30', '18:00'),
        'G': ('18:15', '19:45'),
        'H': ('20:00', '21:30')
    }
    start_time, end_time = time_mapping[cell.column_letter]  # Получаем начальное и конечное время на основе буквы столбца
    return start_time, end_time

# Функция для формирования информации о занятии
def form_exercise(date, cell, next_cell, joined, room, subject, lesson_type, group):
    time_start, time_end = get_time_period(cell)  # Получаем начальное и конечное время занятия
    # Создаём словарь с деталями занятия
    exercise_details = {
        "date": date,
        "time_start": time_start,
        "time_end": time_end,
        "group": group,
        "room": room,
        "type": lesson_type,
        "subject": subject,
        "joined": joined,
    }
    if joined:  # Если занятие объединено со следующим
        time_start_s, time_end_s = get_time_period(next_cell)  # Получаем временной период следующего занятия
        # Добавляем информацию о времени начала и окончания следующего занятия
        exercise_details["time_start_s"] = time_start_s
        exercise_details["time_end_s"] = time_end_s
    return exercise_details  # Возвращаем словарь с информацией о занятии


# Функция для формирования информации о занятии
def form_exercise_student(date, cell, next_cell, joined, room, subject, lesson_type, professor):
    time_start, time_end = get_time_period(cell)  # Получаем начальное и конечное время занятия
    # Создаём словарь с деталями занятия
    exercise_details = {
        "date": date,
        "time_start": time_start,
        "time_end": time_end,
        "professor": professor,
        "room": room,
        "type": lesson_type,
        "subject": subject,
        "joined": joined,
    }
    if joined:  # Если занятие объединено со следующим
        time_start_s, time_end_s = get_time_period(next_cell)  # Получаем временной период следующего занятия
        # Добавляем информацию о времени начала и окончания следующего занятия
        exercise_details["time_start_s"] = time_start_s
        exercise_details["time_end_s"] = time_end_s
    return exercise_details  # Возвращаем словарь с информацией о занятии


def read_professor(file_name):
    # Словарь для сопоставления номеров строк с днями недели
    day_mapping = {
        5: 0,  # понедельник
        6: 0,  # понедельник
        7: 1,  # вторник
        8: 1,  # вторник
        9: 2,  # среда
        10: 2, # среда
        11: 3, # четверг
        12: 3, # четверг
        13: 4, # пятница
        14: 4, # пятница
        15: 5, # суббота
        16: 5  # суббота
    }

    # Загрузка рабочей книги Excel и активного листа
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    lessons = [] # Список для хранения занятий
    added_exercises = [] # Список для отслеживания уже добавленных занятий

    # Извлечение ФИО преподавателя
    name = extract_fio(sheet['C2'].value)

    # Словарь для хранения занятий по датам
    exercises_by_date = {}
    # Итерация по строкам и столбцам листа
    for row in sheet.iter_rows(min_row=5, max_row=16, min_col=2, max_col=8):
        for cell in row:
            if cell.coordinate: # Если ячейка имеет координату
                if cell.value: # Если в ячейке есть значение
                    # Проверка, объединена ли ячейка
                    is_merged = any([cell.coordinate in item for item in sheet.merged_cells.ranges])
                    if is_merged:
                        cell_height = 2 # Высота объединенной ячейки
                    else:
                        cell_height = 1 # Высота необъединенной ячейки

                    every_week = cell_height > 1 # Занятие каждую неделю, если ячейка объединена

                    # Извлечение информации о занятии из ячейки
                    for exercise in extract_professor_info(cell.value):
                        if exercise in added_exercises:
                            continue # Пропускаем, если занятие уже обработано
                        added_exercises.append(exercise)
                        room, subject, lesson_type, group, date_list = exercise
                        # Разбор дат начала и конца занятий
                        for date_range in date_list:
                            start_date_str, end_date_str = date_range.split('-')
                            start_date = datetime.strptime(start_date_str + "." + str(YEAR), '%d.%m.%Y')
                            end_date = datetime.strptime(end_date_str + "." + str(YEAR), '%d.%m.%Y')
                            once_in_two_weeks = True # Переключатель для занятий через неделю
                            for date in get_dates_between(start_date, end_date, day_mapping[cell.row]):
                                # Проверяем, соответствует ли дата текущей строке расписания
                                if date.weekday() == day_mapping[cell.row]:
                                    # Получаем следующую ячейку для проверки, объединено ли занятие
                                    next_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                                    joined = check_if_exercise_joined(exercise, next_cell)
                                    if every_week:
                                        # Формируем структуру занятия, если оно каждую неделю
                                        exercise_struct = form_exercise(date, cell, next_cell, joined, room, subject, lesson_type, group)
                                        # Проверяем и добавляем занятие в словарь по датам и времени
                                        if (date, exercise_struct["time_start"], exercise_struct["time_end"]) not in exercises_by_date:
                                            exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])] = []
                                        exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])].append(exercise)
                                        # Если занятие объединено, добавляем информацию о времени объединенного занятия
                                        if exercise_struct["joined"]:
                                            if (date, exercise_struct["time_start_s"], exercise_struct["time_end_s"]) not in exercises_by_date:
                                                exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])] = []
                                            exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])].append(exercise)
                                        # Добавляем структуру занятия в список занятий
                                        lessons.append(exercise_struct)
                                    else:
                                        # Если занятие через неделю, переключаемся на следующую неделю
                                        if once_in_two_weeks:
                                            exercise_struct = form_exercise(date, cell, next_cell, joined, room, subject, lesson_type, group)
                                            if (date, exercise_struct["time_start"], exercise_struct["time_end"]) not in exercises_by_date:
                                                exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])] = []
                                            exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])].append(exercise)
                                            if exercise_struct["joined"]:
                                                if (date, exercise_struct["time_start_s"], exercise_struct["time_end_s"]) not in exercises_by_date:
                                                    exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])] = []
                                                exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])].append(exercise)
                                            lessons.append(exercise_struct)
                                            once_in_two_weeks = False # Переключаемся на другую неделю
                                        else:
                                            once_in_two_weeks = True # Следующая итерация будет для "другой" недели

    # Строка для сбора ошибок в расписании
    erorrs = ""
    for key, value in exercises_by_date.items():
        if (len(value) > 1):
            # Формирование сообщений об ошибках для перекрывающихся занятий
            erorrs += f'{key[0].strftime("%d.%m")} с {key[1]} до {key[2]} накладываются занятия:\n'
            for val in value:
                erorrs += f"{val[1]} ({val[2]}) у группы {val[3][0]}\n"
    # Возвращаем ФИО преподавателя, список занятий и ошибки
    return name, lessons, erorrs


def read_student(file_name):
    # Словарь для сопоставления номеров строк с днями недели
    day_mapping = {
        5: 0,  # понедельник
        6: 0,  # понедельник
        7: 1,  # вторник
        8: 1,  # вторник
        9: 2,  # среда
        10: 2, # среда
        11: 3, # четверг
        12: 3, # четверг
        13: 4, # пятница
        14: 4, # пятница
        15: 5, # суббота
        16: 5  # суббота
    }

    # Загрузка рабочей книги Excel и активного листа
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    lessons = [] # Список для хранения занятий
    added_exercises = [] # Список для отслеживания уже добавленных занятий

    # Извлечение ФИО преподавателя
    group = extract_group(sheet['C2'].value)

    # Словарь для хранения занятий по датам
    exercises_by_date = {}
    # Итерация по строкам и столбцам листа
    for row in sheet.iter_rows(min_row=5, max_row=16, min_col=2, max_col=8):
        for cell in row:
            if cell.coordinate: # Если ячейка имеет координату
                if cell.value: # Если в ячейке есть значение
                    # Проверка, объединена ли ячейка
                    is_merged = any([cell.coordinate in item for item in sheet.merged_cells.ranges])
                    if is_merged:
                        cell_height = 2 # Высота объединенной ячейки
                    else:
                        cell_height = 1 # Высота необъединенной ячейки

                    every_week = cell_height > 1 # Занятие каждую неделю, если ячейка объединена

                    # Извлечение информации о занятии из ячейки

                    for exercise in extract_student_info(cell.value):
                        if exercise in added_exercises:
                            continue # Пропускаем, если занятие уже обработано
                        added_exercises.append(exercise)
                        room, subject, lesson_type, professor, date_range = exercise
                        # Разбор дат начала и конца занятий
                        start_date_str, end_date_str = date_range.split('-')
                        start_date = datetime.strptime(start_date_str + "." + str(YEAR), '%d.%m.%Y')
                        end_date = datetime.strptime(end_date_str + "." + str(YEAR), '%d.%m.%Y')
                        once_in_two_weeks = True # Переключатель для занятий через неделю
                        for date in get_dates_between(start_date, end_date, day_mapping[cell.row]):
                            # Проверяем, соответствует ли дата текущей строке расписания
                            if date.weekday() == day_mapping[cell.row]:
                                # Получаем следующую ячейку для проверки, объединено ли занятие
                                next_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                                joined = check_if_exercise_joined(exercise, next_cell)
                                if every_week:
                                    # Формируем структуру занятия, если оно каждую неделю
                                    exercise_struct = form_exercise_student(date, cell, next_cell, joined, room, subject, lesson_type, professor)
                                    # Проверяем и добавляем занятие в словарь по датам и времени
                                    if (date, exercise_struct["time_start"], exercise_struct["time_end"]) not in exercises_by_date:
                                        exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])] = []
                                    exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])].append(exercise)
                                    # Если занятие объединено, добавляем информацию о времени объединенного занятия
                                    if exercise_struct["joined"]:
                                        if (date, exercise_struct["time_start_s"], exercise_struct["time_end_s"]) not in exercises_by_date:
                                            exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])] = []
                                        exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])].append(exercise)
                                    # Добавляем структуру занятия в список занятий
                                    lessons.append(exercise_struct)
                                else:
                                    # Если занятие через неделю, переключаемся на следующую неделю
                                    if once_in_two_weeks:
                                        exercise_struct = form_exercise_student(date, cell, next_cell, joined, room, subject, lesson_type, group)
                                        if (date, exercise_struct["time_start"], exercise_struct["time_end"]) not in exercises_by_date:
                                            exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])] = []
                                        exercises_by_date[(date, exercise_struct["time_start"], exercise_struct["time_end"])].append(exercise)
                                        if exercise_struct["joined"]:
                                            if (date, exercise_struct["time_start_s"], exercise_struct["time_end_s"]) not in exercises_by_date:
                                                exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])] = []
                                            exercises_by_date[(date, exercise_struct["time_start_s"], exercise_struct["time_end_s"])].append(exercise)
                                        lessons.append(exercise_struct)
                                        once_in_two_weeks = False # Переключаемся на другую неделю
                                    else:
                                        once_in_two_weeks = True # Следующая итерация будет для "другой" недели
    # Строка для сбора ошибок в расписании
    erorrs = ""
    for key, value in exercises_by_date.items():
        if (len(value) > 1):
            # Формирование сообщений об ошибках для перекрывающихся занятий
            erorrs += f'{key[0].strftime("%d.%m")} с {key[1]} до {key[2]} накладываются занятия:\n'
            for val in value:
                erorrs += f"{val[1]} ({val[0]}) {val[3]}\n"
    # Возвращаем ФИО преподавателя, список занятий и ошибки
    return group, lessons, erorrs


if __name__ == '__main__':
    import os
    import sys

    # Получаем абсолютный путь к директории, где находится main.py
    dir_path = os.path.dirname(os.path.realpath(__file__))
    # Добавляем путь к папке scripts, чтобы мы могли импортировать из неё модули
    scripts_path = os.path.join(dir_path, '..')
    sys.path.append(scripts_path)

    file_name = "data/example.xlsx"
    name, exercises, errors = read_professor(file_name)
    print(exercises)